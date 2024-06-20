from flask import Flask, request, jsonify
import pandas as pd
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/webhook', methods=['POST'])
def webhook():
    data = request.json
    # Extract message details from the JSON data
    message = data['messages'][0]['text']['body']
    sender = data['contacts'][0]['profile']['name']
    # Save the message to Excel
    save_to_excel(sender, message)
    return jsonify({'status': 'success'}), 200

def save_to_excel(sender, message):
    # Define the file path and sheet name
    file_path = 'messages.xlsx'
    sheet_name = 'Sheet1'
    
    # Load the existing workbook
    try:
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        writer.book = book
    except FileNotFoundError:
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
    
    # Create a new DataFrame with the message data
    df = pd.DataFrame([[sender, message]], columns=['Sender', 'Message'])
    
    # Append the data to the existing sheet
    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=writer.sheets[sheet_name].max_row if sheet_name in writer.sheets else 0)
    
    # Save the workbook
    writer.save()

if __name__ == '__main__':
    app.run(port=5000)

